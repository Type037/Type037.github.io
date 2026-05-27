import os
import json
import glob
from typing import List, Dict, Any, Optional
# 需要先安装 'openpyxl' 库：pip install openpyxl
import openpyxl  # type: ignore

# --- 配置参数（对应C#版本的变量）---
# 注意：使用相对于脚本运行位置的相对路径，或绝对路径
EXCEL_DIRECTORY = "mine/data/Excel"  # 存放Excel文件的文件夹相对路径
JSON_OUTPUT_DIRECTORY = "mine/data/Json"  # 存放JSON输出的文件夹相对路径
FIELD_ROW = 1  # 包含字段名的行号（从1开始计数）
DATA_START_ROW = 5  # 数据开始的行号（从1开始计数）
IGNORE_COLUMNS_STR = ""  # 要忽略的列号（从1开始计数），用分号分隔（例如："1;3"）


# -----------------------------------

def parse_ignore_columns(input_str: str) -> List[int]:
    """
    将以分号分隔的忽略列字符串解析为整数列表

    参数:
        input_str: 分号分隔的列号字符串

    返回:
        整数列表，包含要忽略的列号
    """
    if not input_str:
        return []
    try:
        # 转换为整数列表，后续可用于快速查找
        return [int(col.strip()) for col in input_str.split(';') if col.strip()]
    except ValueError:
        print(f"错误：IGNORE_COLUMNS_STR格式无效：'{input_str}'。请使用分号分隔的数字（例如：'1;3'）。")
        return []  # 或者抛出异常


def convert_excel_files_to_json(
        excel_dir: str,
        json_out_dir: str,
        field_row_num: int,
        data_start_row_num: int,
        ignore_cols: List[int]
):
    """
    将excel_dir目录中所有.xlsx文件转换为JSON文件，保存到json_out_dir目录

    参数:
        excel_dir: Excel文件所在目录
        json_out_dir: JSON文件输出目录
        field_row_num: 字段名所在行号
        data_start_row_num: 数据开始的行号
        ignore_cols: 需要忽略的列号列表
    """
    # 获取Excel目录和JSON输出目录的绝对路径
    excel_full_path = os.path.abspath(excel_dir)
    json_output_full_path = os.path.abspath(json_out_dir)

    # 检查Excel目录是否存在
    if not os.path.isdir(excel_full_path):
        print(f"错误：未找到Excel目录：{excel_full_path}")
        return

    # 检查JSON输出目录是否存在，不存在则创建
    if not os.path.isdir(json_output_full_path):
        print(f"创建JSON输出目录：{json_output_full_path}")
        os.makedirs(json_output_full_path, exist_ok=True)  # 若目录不存在则创建

    # 查找目录中所有.xlsx文件
    excel_files = glob.glob(os.path.join(excel_full_path, "*.xlsx"))

    if not excel_files:
        print(f"在{excel_full_path}中未找到.xlsx文件")
        return

    print(f"找到{len(excel_files)}个Excel文件待处理...")

    # 将忽略列转换为集合，提高查找效率
    ignore_cols_set = set(ignore_cols)

    # 遍历处理每个Excel文件
    for excel_file in excel_files:
        file_name = os.path.basename(excel_file)
        print(f"\n正在处理'{file_name}'...")
        try:
            # 加载工作簿，只读取单元格值（不读取公式）
            workbook = openpyxl.load_workbook(excel_file, data_only=True)

            # 检查是否有工作表
            if not workbook.sheetnames:
                print(f"警告：'{file_name}'中未找到工作表。跳过。")
                continue

            # 获取第一个工作表
            worksheet = workbook[workbook.sheetnames[0]]

            # 获取最大列数和最大行数
            max_col = worksheet.max_column
            max_row = worksheet.max_row

            # 检查字段行是否超出最大行数
            if field_row_num > max_row:
                print(f"警告：'{file_name}'中的字段行({field_row_num})超出最大行数({max_row})。跳过。")
                continue

            # --- 读取字段名 ---
            # 映射列索引（从1开始）到字段名
            fields_map: Dict[int, str] = {}
            # 存储我们将要读取数据的列的1-based索引
            valid_field_indices: List[int] = []

            for col in range(1, max_col + 1):
                # 跳过需要忽略的列
                if col not in ignore_cols_set:
                    # 获取字段名单元格的值
                    cell_value = worksheet.cell(row=field_row_num, column=col).value
                    # 处理空值，转换为字符串并去除首尾空格
                    field_name = str(cell_value).strip() if cell_value is not None else ""

                    # 只包含有字段名的列
                    if field_name:
                        fields_map[col] = field_name
                        valid_field_indices.append(col)
                    else:
                        print(f"信息：忽略'{file_name}'中的第{col}列，因为第{field_row_num}行的字段名为空。")

            # 如果没有有效字段名，跳过该文件
            if not fields_map:
                print(f"警告：在'{file_name}'的第{field_row_num}行中未找到有效字段名（已排除指定忽略的列和空单元格）。跳过。")
                continue

            # --- 读取数据行 ---
            data_list: List[Dict[str, Any]] = []

            # 检查数据开始行是否超出最大行数
            if data_start_row_num > max_row:
                print(f"信息：'{file_name}'中的数据开始行({data_start_row_num})超出最大行数({max_row})。不会读取任何数据。")

            # 遍历每一行数据
            for row in range(data_start_row_num, max_row + 1):
                item: Dict[str, Any] = {}
                is_row_valid = False  # 跟踪此行是否至少有一个非空值

                # 读取有效字段对应的列数据
                for col_idx in valid_field_indices:
                    cell_value = worksheet.cell(row=row, column=col_idx).value
                    # 转换值为字符串，处理空值
                    # 这里可以根据需要修改类型处理（例如保留数字、布尔值等类型）
                    value_str = str(cell_value) if cell_value is not None else ""
                    item[fields_map[col_idx]] = value_str

                    # 如果有任何一个单元格有值，认为此行有效
                    if cell_value is not None:
                        is_row_valid = True

                # 只添加有效的行（至少有一个非空值）
                if is_row_valid:
                    data_list.append(item)

            # --- 序列化并写入JSON文件 ---
            # 转换为JSON字符串，缩进4空格美化输出，确保非ASCII字符正常显示
            json_output = json.dumps(data_list, indent=4, ensure_ascii=False)

            # 生成JSON文件名（与Excel文件名相同，后缀改为.json）
            json_file_name = os.path.splitext(file_name)[0] + ".json"
            json_file_path = os.path.join(json_output_full_path, json_file_name)

            try:
                # 写入JSON文件，使用UTF-8编码
                with open(json_file_path, 'w', encoding='utf-8') as f:
                    f.write(json_output)
                print(f"成功将'{file_name}'转换为'{json_file_name}'")
            except IOError as e:
                print(f"写入JSON文件'{json_file_path}'时出错：{e}")

        except Exception as e:
            print(f"处理文件'{file_name}'时出错：{e}")
        finally:
            # 确保工作簿被关闭（释放资源）
            if 'workbook' in locals() and workbook:
                workbook.close()


# --- 主执行块 ---
if __name__ == "__main__":
    print("开始Excel到JSON的转换...")
    # 解析需要忽略的列
    ignored_columns = parse_ignore_columns(IGNORE_COLUMNS_STR)
    # 执行转换
    convert_excel_files_to_json(
        EXCEL_DIRECTORY,
        JSON_OUTPUT_DIRECTORY,
        FIELD_ROW,
        DATA_START_ROW,
        ignored_columns
    )
    print("\n转换过程完成。")
